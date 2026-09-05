import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from django.db import transaction
from django.utils import timezone
from django.http import HttpResponse
from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action, api_view, permission_classes, renderer_classes
from rest_framework.response import Response
from rest_framework.renderers import BaseRenderer, JSONRenderer
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser

from .models import Empleado, RegistroAsistencia, BitacoraAccion, DiaFeriado, AutorizacionHorasExtra, AlertaAsistencia, PermisoAusencia, CompensacionHoras
from .serializers import EmpleadoSerializer, RegistroAsistenciaSerializer, BitacoraAccionSerializer, DiaFeriadoSerializer, AutorizacionHorasExtraSerializer, AlertaAsistenciaSerializer, PermisoAusenciaSerializer, CompensacionHorasSerializer


class ExcelBinaryRenderer(BaseRenderer):
    media_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    format = 'xlsx'
    charset = None
    render_style = 'binary'

    def render(self, data, accepted_media_type=None, renderer_context=None):
        return data


def _get_clean_ip(request):
    forwarded = request.META.get('HTTP_X_FORWARDED_FOR')
    if forwarded:
        ip = forwarded.split(',')[0].strip()
    else:
        ip = request.META.get('REMOTE_ADDR', '')
    return ip[:45]


class EmpleadoViewSet(viewsets.ModelViewSet):
    queryset = Empleado.objects.all()
    serializer_class = EmpleadoSerializer
    permission_classes = [permissions.AllowAny]

    def perform_create(self, serializer):
        emp = serializer.save()
        BitacoraAccion.objects.create(
            usuario=self.request.user if self.request.user.is_authenticated else None,
            accion='CREAR_EMPLEADO',
            descripcion=f"Empleado '{emp.nombre} {emp.apellido}' ({emp.get_cargo_display()}) creado.",
            ip_address=_get_clean_ip(self.request)
        )

    def perform_update(self, serializer):
        emp = serializer.save()
        BitacoraAccion.objects.create(
            usuario=self.request.user if self.request.user.is_authenticated else None,
            accion='EDITAR_EMPLEADO',
            descripcion=f"Empleado #{emp.id} '{emp.nombre} {emp.apellido}' actualizado.",
            ip_address=_get_clean_ip(self.request)
        )

    def perform_destroy(self, instance):
        nombre = f"{instance.nombre} {instance.apellido}"
        instance.delete()
        BitacoraAccion.objects.create(
            usuario=self.request.user if self.request.user.is_authenticated else None,
            accion='ELIMINAR_EMPLEADO',
            descripcion=f"Empleado '{nombre}' eliminado del sistema.",
            ip_address=_get_clean_ip(self.request)
        )

    @action(detail=True, methods=['post'], url_path='regenerar-qr')
    def regenerar_qr(self, request, pk=None):
        import uuid
        empleado = self.get_object()
        empleado.qr_code_token = uuid.uuid4()
        empleado.save()
        BitacoraAccion.objects.create(
            usuario=self.request.user if self.request.user.is_authenticated else None,
            accion='EDITAR_EMPLEADO',
            descripcion=f"Código QR del empleado #{empleado.id} '{empleado.nombre} {empleado.apellido}' regenerado.",
            ip_address=_get_clean_ip(self.request)
        )
        return Response({'status': 'ok', 'qr_code_token': str(empleado.qr_code_token)})


class RegistroAsistenciaViewSet(viewsets.ModelViewSet):
    queryset = RegistroAsistencia.objects.all().select_related('empleado')
    serializer_class = RegistroAsistenciaSerializer
    permission_classes = [permissions.AllowAny]
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def perform_destroy(self, instance):
        empleado = instance.empleado
        fecha_instancia = instance.fecha_hora.astimezone(timezone.get_current_timezone()).date()

        if instance.foto_verificacion:
            try:
                instance.foto_verificacion.delete(save=False)
            except Exception:
                pass
        emp_nombre = f"{empleado.nombre} {empleado.apellido}"
        evento_str = instance.get_tipo_evento_display()
        fecha_str = instance.fecha_hora.strftime("%d/%m/%Y %I:%M %p")

        instance.delete()

        # Recalcular automáticamente saldo de deuda para revertir cualquier déficit huérfano
        try:
            _recalcular_horas_pendientes_empleado(empleado, fecha_instancia)
        except Exception:
            pass

        try:
            BitacoraAccion.objects.create(
                usuario=self.request.user if self.request.user.is_authenticated else None,
                accion='REGISTRO_MANUAL',
                descripcion=f"Marcaje de {evento_str} ({fecha_str}) para {emp_nombre} eliminado por el administrador por corrección.",
                ip_address=_get_clean_ip(self.request)
            )
        except Exception:
            pass

    def perform_update(self, serializer):
        empleado_anterior = serializer.instance.empleado
        instance = serializer.save()
        try:
            fecha_instancia = instance.fecha_hora.astimezone(timezone.get_current_timezone()).date()
            _recalcular_horas_pendientes_empleado(instance.empleado, fecha_instancia)
            if empleado_anterior and empleado_anterior.id != instance.empleado.id:
                _recalcular_horas_pendientes_empleado(empleado_anterior, fecha_instancia)
        except Exception:
            pass

        try:
            BitacoraAccion.objects.create(
                usuario=self.request.user if self.request.user.is_authenticated else None,
                accion='REGISTRO_MANUAL',
                descripcion=(
                    f"Marcaje #{instance.id} corregido por administrador: "
                    f"Asignado a {instance.empleado.nombre} {instance.empleado.apellido} ({instance.get_tipo_evento_display()})."
                ),
                ip_address=_get_clean_ip(self.request)
            )
        except Exception:
            pass

    def perform_create(self, serializer):
        instance = serializer.save()
        try:
            fecha_instancia = instance.fecha_hora.astimezone(timezone.get_current_timezone()).date()
            _recalcular_horas_pendientes_empleado(instance.empleado, fecha_instancia)
        except Exception:
            pass

    @action(detail=False, methods=['post'], url_path='purgar-dia-cero')
    def purgar_dia_cero(self, request):
        pin = str(request.data.get('pin', '')).strip()
        if pin != '2322':
            return Response(
                {'status': 'error', 'mensaje': 'PIN de Gerencia incorrecto. No autorizado.'},
                status=status.HTTP_403_FORBIDDEN
            )

        with transaction.atomic():
            # 1. Liberar fotos y borrar asistencias
            registros = RegistroAsistencia.objects.all()
            for r in registros:
                try:
                    if r.foto_verificacion:
                        r.foto_verificacion.delete(save=False)
                except Exception:
                    pass
            reg_count = registros.delete()[0]

            # 2. Borrar alertas
            alerta_count = AlertaAsistencia.objects.all().delete()[0]

            # 3. Borrar horas extra de prueba
            extra_count = AutorizacionHorasExtra.objects.all().delete()[0]

            # 4. Borrar permisos de prueba
            perm_count = PermisoAusencia.objects.all().delete()[0]

            # 5. Resetear horas_pendientes a 0.00 para todos los empleados
            hoy = timezone.localdate()
            Empleado.objects.all().update(
                horas_pendientes=0.00,
                periodo_horas_pendientes=hoy.replace(day=1)
            )

            # 6. Registrar en bitácora
            BitacoraAccion.objects.create(
                usuario=request.user if request.user.is_authenticated else None,
                accion='CONFIGURACION_SISTEMA',
                descripcion=(
                    f"Purga oficial Día 0 ejecutada con PIN 2322: "
                    f"{reg_count} asistencias, {alerta_count} alertas y {extra_count} horas extra eliminadas. "
                    f"Balances reseteados a 0.00 hrs."
                ),
                ip_address=_get_clean_ip(request)
            )

        return Response({
            'status': 'ok',
            'mensaje': (
                f"Limpieza de Día 0 exitosa: {reg_count} marcajes eliminados, "
                f"{alerta_count} alertas limpiadas y todos los colaboradores en 0.00 hrs."
            )
        })


class BitacoraViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = BitacoraAccion.objects.all()
    serializer_class = BitacoraAccionSerializer
    permission_classes = [permissions.AllowAny]


class DiaFeriadoViewSet(viewsets.ModelViewSet):
    queryset = DiaFeriado.objects.all()
    serializer_class = DiaFeriadoSerializer
    permission_classes = [permissions.AllowAny]


class AutorizacionHorasExtraViewSet(viewsets.ModelViewSet):
    queryset = AutorizacionHorasExtra.objects.all().select_related('empleado')
    serializer_class = AutorizacionHorasExtraSerializer
    permission_classes = [permissions.AllowAny]

    def perform_create(self, serializer):
        autorizacion = serializer.save()
        if autorizacion.estado == 'APROBADO':
            self._procesar_amortizacion_si_aprobado(autorizacion)

    def perform_update(self, serializer):
        instancia_previa = self.get_object()
        estado_previo = instancia_previa.estado

        autorizacion = serializer.save()

        # Si pasa a APROBADO desde otro estado (ej. de PENDIENTE a APROBADO)
        if autorizacion.estado == 'APROBADO' and estado_previo != 'APROBADO':
            self._procesar_amortizacion_si_aprobado(autorizacion)

    def _procesar_amortizacion_si_aprobado(self, autorizacion):
        empleado = autorizacion.empleado
        horas_aprobadas = float(autorizacion.horas_extra_autorizadas or 0.0)
        if horas_aprobadas <= 0:
            horas_aprobadas = float(autorizacion.horas_extra_solicitadas or 0.0)

        deuda_emp = float(empleado.horas_pendientes or 0.0)

        if deuda_emp > 0 and horas_aprobadas > 0:
            res = _aplicar_amortizacion_deuda_empleado(
                empleado=empleado,
                fecha_referencia=autorizacion.fecha,
                horas_a_amortizar=horas_aprobadas,
                request=self.request
            )
            horas_deducidas = res['horas_amortizadas']
            remanente = res['remanente']

            autorizacion.horas_extra_autorizadas = remanente
            nota_comp = f"[Bolsa de Horas: {horas_deducidas:.2f}h amortizadas a deuda previa (saldo: {res['deuda_restante']:.2f}h). {remanente:.2f}h enviadas a nómina]"
            if autorizacion.comentario:
                autorizacion.comentario = f"{autorizacion.comentario} | {nota_comp}"
            else:
                autorizacion.comentario = nota_comp

            autorizacion.save(update_fields=['horas_extra_autorizadas', 'comentario', 'updated_at'])


class AlertaAsistenciaViewSet(viewsets.ModelViewSet):
    queryset = AlertaAsistencia.objects.all().select_related('empleado')
    serializer_class = AlertaAsistenciaSerializer
    permission_classes = [permissions.AllowAny]

    def list(self, request, *args, **kwargs):
        # Al listar alertas, verificar ausencias de la semana y mantenimiento semestral
        self._verificar_ausencias_semanales()
        self._verificar_mantenimiento_semestral()
        return super().list(request, *args, **kwargs)

    def _verificar_mantenimiento_semestral(self):
        """
        Verifica si existen registros de asistencia o fotos de hace más de 180 días (6 meses).
        Si existen, genera un recordatorio al administrador para ejecutar la depuración semestral.
        """
        import datetime
        limite = timezone.now() - datetime.timedelta(days=180)
        hay_antiguos = RegistroAsistencia.objects.filter(fecha_hora__lt=limite).exists()
        if hay_antiguos:
            ya_notificado = AlertaAsistencia.objects.filter(
                tipo='MANTENIMIENTO',
                leida=False
            ).exists()
            if not ya_notificado:
                AlertaAsistencia.objects.create(
                    tipo='MANTENIMIENTO',
                    empleado=None,
                    titulo="Mantenimiento Semestral Sugerido",
                    mensaje=(
                        "Existen registros de asistencia y fotografías con más de 6 meses de antigüedad. "
                        "Puede ejecutar la Depuración Semestral desde el Panel para mantener optimizado el almacenamiento gratuito."
                    ),
                    leida=False
                )

    def _verificar_ausencias_semanales(self):
        """
        Revisa la semana en curso (Lunes a hoy). Si un empleado tiene 2 o más días sin marcaje
        y no es feriado, genera la alerta de SEGUNDA_AUSENCIA para que el admin tome una decisión.
        """
        import datetime
        hoy = timezone.localdate()
        inicio_semana = hoy - datetime.timedelta(days=hoy.weekday())
        
        # Obtener feriados de la semana
        feriados = set(DiaFeriado.objects.filter(
            fecha__gte=inicio_semana,
            fecha__lte=hoy
        ).values_list('fecha', flat=True))

        # Si no hay registros de asistencia en la semana (sistema recién estrenado o purgado), no generar falsas alertas
        if not RegistroAsistencia.objects.filter(fecha_hora__date__gte=inicio_semana).exists():
            return

        empleados = Empleado.objects.filter(activo=True)
        for emp in empleados:
            # Si el empleado nunca ha marcado en el sistema, no evaluar ausencias previas a su inicio
            primer_registro = RegistroAsistencia.objects.filter(empleado=emp).order_by('fecha_hora').first()
            if not primer_registro:
                continue

            # Obtener días con marcajes en la semana
            dias_con_marcaje = set(RegistroAsistencia.objects.filter(
                empleado=emp,
                fecha_hora__date__gte=inicio_semana,
                fecha_hora__date__lte=hoy
            ).dates('fecha_hora', 'day'))

            # Obtener días con permiso o vacaciones autorizadas
            permisos_emp = PermisoAusencia.objects.filter(
                empleado=emp,
                fecha_inicio__lte=hoy,
                fecha_fin__gte=inicio_semana
            )
            dias_permiso = set()
            for p in permisos_emp:
                d_curr = max(p.fecha_inicio, inicio_semana)
                d_fin = min(p.fecha_fin, hoy)
                while d_curr <= d_fin:
                    dias_permiso.add(d_curr)
                    d_curr += datetime.timedelta(days=1)

            dias_sin_marcaje = []
            fecha_inicio_eval = max(inicio_semana, primer_registro.fecha_hora.date())
            curr = fecha_inicio_eval
            # Revisar hasta ayer (hoy aún puede marcar durante su turno)
            while curr < hoy:
                if curr not in feriados and curr not in dias_con_marcaje and curr not in dias_permiso:
                    dias_sin_marcaje.append(curr)
                curr += datetime.timedelta(days=1)

            # Si tiene 2 o más días sin marcar en la semana (excluyendo permisos y vacaciones)
            if len(dias_sin_marcaje) >= 2:
                # Verificar si ya existe una alerta de SEGUNDA_AUSENCIA para esta semana
                titulo_busqueda = f"Segunda Ausencia Semanal: {emp.nombre} {emp.apellido}"
                alerta_existente = AlertaAsistencia.objects.filter(
                    empleado=emp,
                    tipo='SEGUNDA_AUSENCIA',
                    created_at__date__gte=inicio_semana
                ).exists()

                if not alerta_existente:
                    fechas_str = ", ".join(d.strftime('%d/%m') for d in dias_sin_marcaje)
                    AlertaAsistencia.objects.create(
                        tipo='SEGUNDA_AUSENCIA',
                        empleado=emp,
                        titulo=titulo_busqueda,
                        mensaje=(
                            f"El empleado {emp.nombre} {emp.apellido} acumula {len(dias_sin_marcaje)} días sin registrar asistencia "
                            f"esta semana ({fechas_str}). El 1er día cuenta como día libre. "
                            f"Decida si autoriza la falta o si suma las 8 horas como deuda pendiente."
                        ),
                        leida=False
                    )

    @action(detail=True, methods=['post'], url_path='resolver')
    def resolver_alerta(self, request, pk=None):
        """
        Resuelve una alerta:
        decision='JUSTIFICAR': Marca como justificada (no genera deuda)
        decision='SUMAR_DEUDA': Suma 8 horas al saldo de horas_pendientes del empleado
        """
        alerta = self.get_object()
        decision = request.data.get('decision', 'JUSTIFICAR')  # 'JUSTIFICAR' o 'SUMAR_DEUDA'
        empleado = alerta.empleado

        if empleado:
            if decision == 'SUMAR_DEUDA' and alerta.tipo == 'SEGUNDA_AUSENCIA':
                hoy = timezone.localdate()
                primer_dia_mes = hoy.replace(day=1)
                if empleado.periodo_horas_pendientes != primer_dia_mes:
                    empleado.horas_pendientes = 0.00
                    empleado.periodo_horas_pendientes = primer_dia_mes

                empleado.horas_pendientes = float(empleado.horas_pendientes) + 8.00
                empleado.save(update_fields=['horas_pendientes', 'periodo_horas_pendientes'])
                
                BitacoraAccion.objects.create(
                    usuario=request.user if request.user.is_authenticated else None,
                    accion='REGISTRO_MANUAL',
                    descripcion=f"Se sumaron 8.0 hrs de deuda a {empleado.nombre} {empleado.apellido} por ausencia no justificada (Alerta #{alerta.id}).",
                    ip_address=_get_clean_ip(request)
                )
            else:
                BitacoraAccion.objects.create(
                    usuario=request.user if request.user.is_authenticated else None,
                    accion='REGISTRO_MANUAL',
                    descripcion=f"Se justificó la alerta #{alerta.id} de {empleado.nombre} {empleado.apellido} (sin recargo de horas).",
                    ip_address=_get_clean_ip(request)
                )

        alerta.leida = True
        alerta.save(update_fields=['leida'])
        return Response({
            'status': 'ok',
            'mensaje': 'Alerta procesada correctamente.',
            'empleado_horas_pendientes': float(empleado.horas_pendientes) if empleado else 0.0
        })

    @action(detail=False, methods=['post'], url_path='marcar-todas-leidas')
    def marcar_todas_leidas(self, request):
        AlertaAsistencia.objects.filter(leida=False).update(leida=True)
        return Response({'status': 'ok', 'mensaje': 'Todas las alertas han sido marcadas como leídas.'})

    @action(detail=False, methods=['post'], url_path='limpiar-leidas')
    def limpiar_leidas(self, request):
        """
        Elimina permanentemente todas las alertas que ya fueron leídas/resueltas.
        Mantiene intactas las alertas pendientes (leida=False).
        """
        eliminadas, _ = AlertaAsistencia.objects.filter(leida=True).delete()
        BitacoraAccion.objects.create(
            usuario=request.user if request.user.is_authenticated else None,
            accion='MANTENIMIENTO_DEPURACION',
            descripcion=f"Historial de notificaciones leídas limpiado manualmente. {eliminadas} alerta(s) eliminada(s).",
            ip_address=_get_clean_ip(request)
        )
        return Response({
            'status': 'ok',
            'mensaje': f'Se eliminaron {eliminadas} notificación(es) leída(s) del historial.',
            'eliminadas': eliminadas,
        })

    @action(detail=False, methods=['post'], url_path='depurar-historial')
    def depurar_historial(self, request):
        """
        Depura fotografías y datos prescindibles con más de 6 meses (180 días) de antigüedad,
        manteniendo intactos a los empleados, sus carnets y sus saldos de horas.
        """
        import datetime
        meses = int(request.data.get('meses', 6))
        limite_fecha = timezone.now() - datetime.timedelta(days=meses * 30)

        # 1. Liberar fotografías de registros antiguos (tanto en disco como en Supabase)
        from django.db.models import Q
        registros_con_foto = RegistroAsistencia.objects.filter(
            fecha_hora__lt=limite_fecha
        ).filter(
            Q(foto_verificacion__isnull=False) | Q(foto_base64__isnull=False)
        )
        
        fotos_liberadas = 0
        for reg in registros_con_foto:
            if reg.foto_verificacion:
                try:
                    reg.foto_verificacion.delete(save=False)
                except Exception:
                    pass
                reg.foto_verificacion = None
            reg.foto_base64 = None
            reg.save(update_fields=['foto_verificacion', 'foto_base64'])
            fotos_liberadas += 1

        # 2. Limpiar bitácora antigua
        bitacora_eliminada = BitacoraAccion.objects.filter(fecha_hora__lt=limite_fecha).delete()[0]

        # 3. Limpiar alertas antiguas ya leídas
        alertas_eliminadas = AlertaAsistencia.objects.filter(created_at__lt=limite_fecha, leida=True).delete()[0]

        # Marcar alertas de mantenimiento como leídas
        AlertaAsistencia.objects.filter(tipo='MANTENIMIENTO').update(leida=True)

        BitacoraAccion.objects.create(
            usuario=request.user if request.user.is_authenticated else None,
            accion='MANTENIMIENTO_DEPURACION',
            descripcion=f"Depuración semestral ejecutada: {fotos_liberadas} fotos liberadas, {bitacora_eliminada} logs de bitácora y {alertas_eliminadas} alertas antiguas depuradas.",
            ip_address=_get_clean_ip(request)
        )

        return Response({
            'status': 'ok',
            'mensaje': f'Depuración completada exitosamente. Se liberaron {fotos_liberadas} fotografías y se limpiaron {bitacora_eliminada + alertas_eliminadas} registros antiguos.',
            'fotos_liberadas': fotos_liberadas,
            'bitacora_eliminada': bitacora_eliminada,
            'alertas_eliminadas': alertas_eliminadas,
        })


class PermisoAusenciaViewSet(viewsets.ModelViewSet):
    queryset = PermisoAusencia.objects.all().select_related('empleado')
    serializer_class = PermisoAusenciaSerializer
    permission_classes = [permissions.AllowAny]

    def perform_create(self, serializer):
        permiso = serializer.save()
        BitacoraAccion.objects.create(
            usuario=self.request.user if self.request.user.is_authenticated else None,
            accion='CREAR_PERMISO',
            descripcion=f"Permiso/Vacaciones registrado para {permiso.empleado.nombre} {permiso.empleado.apellido}: {permiso.get_tipo_display()} ({permiso.fecha_inicio} a {permiso.fecha_fin}).",
            ip_address=_get_clean_ip(self.request)
        )

    def perform_destroy(self, instance):
        desc = f"Permiso cancelado para {instance.empleado.nombre} {instance.empleado.apellido}: {instance.get_tipo_display()} ({instance.fecha_inicio} a {instance.fecha_fin})."
        instance.delete()
        BitacoraAccion.objects.create(
            usuario=self.request.user if self.request.user.is_authenticated else None,
            accion='ELIMINAR_PERMISO',
            descripcion=desc,
            ip_address=_get_clean_ip(self.request)
        )


class CompensacionHorasViewSet(viewsets.ModelViewSet):
    queryset = CompensacionHoras.objects.select_related('empleado').all()
    serializer_class = CompensacionHorasSerializer
    permission_classes = [permissions.AllowAny]

    def perform_destroy(self, instance):
        empleado = instance.empleado
        fecha_ref = instance.fecha_compensacion
        instance.delete()
        _recalcular_horas_pendientes_empleado(empleado, fecha_ref)


def _procesar_foto_a_base64(foto):
    """
    Comprime la fotografía facial a 320x320 JPEG calidad 55% y genera un Data-URI
    ultra-ligero (~12-15 KB) que se almacena permanentemente en Supabase (PostgreSQL).
    """
    if not foto:
        return None
    try:
        import base64
        import io
        from PIL import Image
        foto.seek(0)
        img = Image.open(foto)
        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')
        img.thumbnail((320, 320), Image.Resampling.LANCZOS)
        buf = io.BytesIO()
        img.save(buf, format='JPEG', quality=55, optimize=True)
        encoded = base64.b64encode(buf.getvalue()).decode('utf-8')
        foto.seek(0)
        return f"data:image/jpeg;base64,{encoded}"
    except Exception as e:
        print(f"Error procesando foto a base64: {e}")
        return None


def _autodetectar_tipo_evento(registros_hoy, fecha_hora_registro):
    """
    Auto-detección inteligente de eventos para El Bodegón:
    - Primer evento del día -> siempre ENTRADA.
    - Cierre nocturno (después de las 9:30 PM / 1290 min) -> siempre SALIDA_DEFINITIVA.
    - Segundo evento:
      * Si han transcurrido >= 6.8 horas (ej: 9am-5pm, o 3pm-11pm) -> SALIDA_DEFINITIVA.
      * Si han transcurrido < 6.8 horas y es mediodía/tarde -> SALIDA_QUEBRADA (Pausa).
    - Tercer evento (después de SALIDA_QUEBRADA) -> ENTRADA_QUEBRADA (Retorno).
    - Cuarto evento (después de ENTRADA_QUEBRADA) -> SALIDA_DEFINITIVA (Cierre).
    """
    if not registros_hoy.exists():
        return 'ENTRADA'

    dt_local = fecha_hora_registro.astimezone(timezone.get_current_timezone())
    hora_mins = dt_local.hour * 60 + dt_local.minute
    cant = registros_hoy.count()
    ultimo = registros_hoy.last()
    primero = registros_hoy.first()

    # Regla de Cierre: Después de las 9:30 PM (21:30 = 1290 min) cualquier salida es definitiva
    if hora_mins >= 1290:
        return 'SALIDA_DEFINITIVA'

    if cant == 1:
        # Segundo evento del día
        segundos = (fecha_hora_registro - primero.fecha_hora).total_seconds()
        horas_transcurridas = segundos / 3600.0
        # Si ya completó jornada corrida (>= 6.8h) es salida definitiva (ej: 9am-5pm)
        if horas_transcurridas >= 6.8:
            return 'SALIDA_DEFINITIVA'
        # Si lleva menos tiempo y estamos entre 12:30 PM y 6:30 PM, es salida a pausa
        if 750 <= hora_mins <= 1140:
            return 'SALIDA_QUEBRADA'
        return 'SALIDA_DEFINITIVA'

    elif cant == 2:
        # Tercer evento del día: si el último fue una salida a pausa (o salida), es retorno de pausa
        if ultimo.tipo_evento in ('SALIDA_QUEBRADA', 'SALIDA_DEFINITIVA'):
            return 'ENTRADA_QUEBRADA'
        return 'SALIDA_DEFINITIVA'

    elif cant == 3:
        # Cuarto evento del día: tras retornar de pausa, el siguiente es salida definitiva
        return 'SALIDA_DEFINITIVA'

    else:
        if ultimo.tipo_evento in ('SALIDA_DEFINITIVA', 'SALIDA_QUEBRADA'):
            return 'ENTRADA'
        return 'SALIDA_DEFINITIVA'


def _evaluar_alertas_asistencia(registro, empleado, registros_actualizados, horas_netas_hoy, es_offline=False):
    """
    Evalúa puntualidad con la regla de gracia de 10 min de El Bodegón:
    - 4 Franjas Base de Entrada: 9:00 AM (540), 11:00 AM (660), 12:00 PM (720), 3:00 PM (900).
    - Margen de gracia: 10 minutos (1 a 10 min retraso leve informativo, NO alerta al admin).
    - Tardanza severa: > 10 minutos (genera alerta en campanita y acta imprimible).
    - Salida definitiva: alerta solo si déficit > 10 minutos de las 8 horas (horas_netas_hoy < 7.83).
    """
    try:
        hora_actual = registro.fecha_hora.astimezone(timezone.get_current_timezone())
        mins_marcados = hora_actual.hour * 60 + hora_actual.minute
        tipo = registro.tipo_evento
        tag_offline = " (Offline)" if es_offline else ""

        alerta_creada = False
        alerta_tipo = ''
        alerta_titulo = ''
        alerta_mensaje = ''

        if tipo == 'ENTRADA':
            # Franjas Base: 9:00 AM (540), 11:00 AM (660), 12:00 PM (720), 3:00 PM (900)
            # Detección inteligente de horario (Mecanismo 1):
            # Si hay retorno de pausa en el día, verificar si fue hacia las 6 PM o 7 PM
            retorno_q = next((r for r in registros_actualizados if r.tipo_evento == 'ENTRADA_QUEBRADA'), None)
            es_quebrado_11am = False
            if retorno_q:
                mins_ret = retorno_q.fecha_hora.astimezone(timezone.get_current_timezone()).hour * 60 + retorno_q.fecha_hora.astimezone(timezone.get_current_timezone()).minute
                if mins_ret >= 1115:  # 6:35 PM o después -> retorno de 7:00 PM (turno 11 AM)
                    es_quebrado_11am = True

            if es_quebrado_11am:
                franja_base = 660  # 11:00 AM
            elif 675 <= mins_marcados <= 735:
                # Entre 11:15 AM y 12:15 PM -> Turno de 12:00 PM (ej. Uriel a las 11:26 AM o Carlos a las 11:55 AM)
                franja_base = 720  # 12:00 PM
            else:
                FRANJAS = [540, 660, 720, 900, 1020]
                franja_base = min(FRANJAS, key=lambda b: abs(mins_marcados - b))

            diff = mins_marcados - franja_base
            HORA_LABELS = {540: '9:00 AM', 660: '11:00 AM', 720: '12:00 PM', 900: '3:00 PM', 1020: '5:00 PM'}
            
            # Solo alertar si excede los 10 minutos de gracia respecto a su turno
            if diff > 10:
                alerta_creada = True
                alerta_tipo = 'TARDANZA'
                alerta_titulo = f"Tardanza severa{tag_offline}: {empleado.nombre} {empleado.apellido}"
                alerta_mensaje = (
                    f"Llegó {diff} min tarde respecto a su turno ({HORA_LABELS[franja_base]}). "
                    f"Marcaje a las {hora_actual.strftime('%I:%M %p')}."
                )

        elif tipo == 'ENTRADA_QUEBRADA':
            # Hora esperada de retorno: 7:00 PM (1140) si el primer bloque fue de ~4 horas (11am a 3pm),
            # o 6:00 PM (1080) si fue de ~3 horas (12pm a 3pm)
            primera_ent = next((r for r in registros_actualizados if r.tipo_evento == 'ENTRADA'), None)
            primera_pausa = next((r for r in registros_actualizados if r.tipo_evento == 'SALIDA_QUEBRADA'), None)
            hora_esperada = 1080  # 6:00 PM por defecto
            if primera_ent and primera_pausa:
                duracion_b1 = (primera_pausa.fecha_hora - primera_ent.fecha_hora).total_seconds() / 3600.0
                if duracion_b1 >= 3.6:
                    hora_esperada = 1140  # 7:00 PM

            diff_ret = mins_marcados - hora_esperada
            ret_label = '7:00 PM' if hora_esperada == 1140 else '6:00 PM'
            if diff_ret > 10:
                alerta_creada = True
                alerta_tipo = 'TARDANZA'
                alerta_titulo = f"Tardanza Retorno de Pausa{tag_offline}: {empleado.nombre} {empleado.apellido}"
                alerta_mensaje = (
                    f"Regresó de la pausa {diff_ret} min tarde respecto a su horario ({ret_label}). "
                    f"Marcaje a las {hora_actual.strftime('%I:%M %p')}."
                )

        elif tipo == 'SALIDA_DEFINITIVA':
            # Detectar si el colaborador inició en el turno vespertino especial de las 5:00 PM (ej. Xiomara Castillo)
            primera_ent = next((r for r in registros_actualizados if r.tipo_evento == 'ENTRADA'), None)
            es_turno_5pm = False
            if primera_ent:
                dt_pe = primera_ent.fecha_hora.astimezone(timezone.get_current_timezone())
                mins_pe = dt_pe.hour * 60 + dt_pe.minute
                if abs(mins_pe - 1020) <= 45:  # Entrada entre 4:15 PM y 5:45 PM
                    es_turno_5pm = True

            if es_turno_5pm:
                # Turno especial de 6 horas acordadas (5:00 PM a 11:00 PM):
                # Solo alertar si se retira con más de 10 min de anticipación a sus 6 horas (< 5.83h)
                if horas_netas_hoy < 5.83:
                    deficit_mins = int(round((6.0 - horas_netas_hoy) * 60))
                    if deficit_mins > 10:
                        alerta_creada = True
                        alerta_tipo = 'SALIDA_ANTICIPADA'
                        alerta_titulo = f"Salida anticipada (Turno 5 PM){tag_offline}: {empleado.nombre} {empleado.apellido}"
                        alerta_mensaje = (
                            f"Se retiró antes de cumplir su jornada acordada de 6 horas. Acumuló {round(horas_netas_hoy, 2)} hrs "
                            f"(Déficit de {deficit_mins} min)."
                        )
            else:
                # Jornada estándar de 8 horas:
                # Solo alertar si el déficit supera los 10 minutos de la jornada de 8 horas (menos de 7.83h)
                if horas_netas_hoy < 7.83:
                    deficit_mins = int(round((8.0 - horas_netas_hoy) * 60))
                    if deficit_mins > 10:
                        alerta_creada = True
                        alerta_tipo = 'SALIDA_ANTICIPADA'
                        alerta_titulo = f"Jornada incompleta{tag_offline}: {empleado.nombre} {empleado.apellido}"
                        alerta_mensaje = (
                            f"Se retiró antes de cumplir sus 8 horas. Acumuló {round(horas_netas_hoy, 2)} hrs "
                            f"(Déficit de {deficit_mins} min)."
                        )

        if alerta_creada:
            AlertaAsistencia.objects.create(
                tipo=alerta_tipo,
                empleado=empleado,
                titulo=alerta_titulo,
                mensaje=alerta_mensaje
            )
    except Exception as ex:
        print(f"Error evaluando alertas: {ex}")


# ==========================================
# 🟢 ENDPOINT PRINCIPAL DEL KIOSCO DE MARCAJE
# ==========================================

@api_view(['POST'])
@permission_classes([permissions.AllowAny])
@renderer_classes([JSONRenderer])
def marcar_asistencia_kiosco(request):
    """
    Endpoint invocado por el Kiosco al escanear un Código QR.
    Acepta: qr_token, tipo_evento (opcional), foto (opcional).
    """
    qr_token = request.data.get('qr_token') or request.data.get('token')
    if not qr_token:
        return Response({'status': 'error', 'mensaje': 'Código QR no proporcionado.'}, status=400)

    qr_token = qr_token.strip()

    # Buscar empleado activo directamente por su token o por el prefijo UUID
    empleado = Empleado.objects.filter(qr_code_token=qr_token, activo=True).first()
    if not empleado and '.' in qr_token:
        raw_uuid = qr_token.split('.', 1)[0]
        empleado = Empleado.objects.filter(qr_code_token__startswith=raw_uuid, activo=True).first()

    if not empleado:
        return Response({'status': 'error', 'mensaje': 'Código QR no encontrado en base de datos o empleado inactivo.'}, status=400)

    # Permite enviar fecha_hora simulada desde el cliente (para testing dinámico)
    fecha_hora_simulada = request.data.get('fecha_hora')
    if fecha_hora_simulada:
        from django.utils.dateparse import parse_datetime
        dt_sim = parse_datetime(fecha_hora_simulada)
        if dt_sim:
            if timezone.is_naive(dt_sim):
                dt_sim = timezone.make_aware(dt_sim, timezone.get_current_timezone())
            dt_local = dt_sim.astimezone(timezone.get_current_timezone())
            fecha_hora_registro = dt_sim
            hoy = dt_local.date()
        else:
            fecha_hora_registro = timezone.now()
            hoy = timezone.localdate()
    else:
        fecha_hora_registro = timezone.now()
        hoy = timezone.localdate()

    # Definir rango de fecha local en UTC para SQLite-safe query
    import datetime
    start_dt = timezone.make_aware(datetime.datetime.combine(hoy, datetime.time.min), timezone.get_current_timezone())
    end_dt = timezone.make_aware(datetime.datetime.combine(hoy, datetime.time.max), timezone.get_current_timezone())

    # Buscar marcajes del empleado en la fecha correspondiente (Hora Nicaragua)
    registros_hoy = RegistroAsistencia.objects.filter(
        empleado=empleado,
        fecha_hora__range=(start_dt, end_dt)
    ).order_by('fecha_hora')

    # ── BLOQUEO ANTI-REBOTE (Cooldown de 5 minutos) ────────────────────────
    if registros_hoy.exists():
        ultimo_reg = registros_hoy.last()
        segundos_diff = (fecha_hora_registro - ultimo_reg.fecha_hora).total_seconds()
        if 0 <= segundos_diff < 300:  # Menos de 5 minutos (300 seg)
            hora_local_str = ultimo_reg.fecha_hora.astimezone(timezone.get_current_timezone()).strftime('%I:%M %p')
            minutos_hace = max(1, int(segundos_diff // 60))
            return Response({
                'status': 'cooldown',
                'mensaje': (
                    f"¡Hola {empleado.nombre}! Tu marcaje de {ultimo_reg.get_tipo_evento_display()} "
                    f"ya fue registrado a las {hora_local_str} (hace {minutos_hace} min). "
                    f"No te preocupes, tu registro está guardado."
                ),
                'registro': RegistroAsistenciaSerializer(ultimo_reg, context={'request': request}).data,
                'horas_trabajadas_hoy': round(_calcular_horas_netas_dia(list(registros_hoy)), 2)
            })

    # Tipo de evento solicitado o auto-detectado inteligentemente por franja horaria
    tipo_evento = request.data.get('tipo_evento')
    dt_local = fecha_hora_registro.astimezone(timezone.get_current_timezone())
    hora_mins = dt_local.hour * 60 + dt_local.minute

    if not tipo_evento:
        tipo_evento = _autodetectar_tipo_evento(registros_hoy, fecha_hora_registro)

    foto = request.FILES.get('foto')
    foto_b64 = _procesar_foto_a_base64(foto)

    # ── DETECCIÓN DE HORARIO INUSUAL / MADRUGADA ───────────────────────────
    if 1 <= dt_local.hour < 10:
        AlertaAsistencia.objects.create(
            tipo='MARCACION_SOSPECHOSA',
            empleado=empleado,
            titulo=f"Marcación fuera de horario: {empleado.nombre} {empleado.apellido}",
            mensaje=(
                f"Se detectó un marcaje a las {dt_local.strftime('%I:%M %p')} "
                f"({tipo_evento}), fuera del horario operativo regular del restaurante."
            ),
            leida=False
        )

    with transaction.atomic():
        registro = RegistroAsistencia.objects.create(
            empleado=empleado,
            tipo_evento=tipo_evento,
            foto_verificacion=foto if foto else None,
            foto_base64=foto_b64,
            fecha_hora=fecha_hora_registro,
            ip_address=_get_clean_ip(request)
        )

    # Calcular horas trabajadas hoy acumuladas hasta el momento
    registros_actualizados = list(RegistroAsistencia.objects.filter(
        empleado=empleado,
        fecha_hora__range=(start_dt, end_dt)
    ).order_by('fecha_hora'))

    horas_netas_hoy = _calcular_horas_netas_dia(registros_actualizados)

    # ── ALERTA: ENTRADA sin SALIDA previa del día anterior ─────────────────
    # Si hoy el empleado marca ENTRADA pero ayer tenía una ENTRADA sin cerrar, alertar al admin
    if tipo_evento == 'ENTRADA':
        ayer = hoy - datetime.timedelta(days=1)
        start_ayer = timezone.make_aware(datetime.datetime.combine(ayer, datetime.time.min), timezone.get_current_timezone())
        end_ayer = timezone.make_aware(datetime.datetime.combine(ayer, datetime.time.max), timezone.get_current_timezone())
        regs_ayer = list(RegistroAsistencia.objects.filter(
            empleado=empleado,
            fecha_hora__range=(start_ayer, end_ayer)
        ).order_by('fecha_hora'))
        if regs_ayer:
            ultimo_ayer = regs_ayer[-1].tipo_evento
            if ultimo_ayer in ('ENTRADA', 'ENTRADA_QUEBRADA'):
                # Ayer quedó con entrada sin salida
                AlertaAsistencia.objects.get_or_create(
                    tipo='REGISTRO_INCOMPLETO',
                    empleado=empleado,
                    titulo=f"Registro incompleto: {empleado.nombre} {empleado.apellido}",
                    defaults={
                        'mensaje': (
                            f"El día {ayer.strftime('%d/%m/%Y')} el empleado registró "
                            f"{regs_ayer[-1].get_tipo_evento_display()} a las "
                            f"{regs_ayer[-1].fecha_hora.astimezone(timezone.get_current_timezone()).strftime('%I:%M %p')} "
                            f"pero nunca registró su Salida. "
                            f"Por favor, agregue la salida manualmente para calcular correctamente sus horas."
                        ),
                        'leida': False
                    }
                )

    # ── BOLSA DE HORAS: Compensación Automática o Solicitud de Horas Extra ──
    comp_info = {
        'horas_amortizadas': 0.0,
        'deuda_restante': float(empleado.horas_pendientes),
        'horas_extra_solicitadas': 0.0,
    }
    if tipo_evento == 'SALIDA_DEFINITIVA':
        comp_info = _procesar_compensacion_y_horas_extra(empleado, hoy, horas_netas_hoy, request)
        # Horas extra por 7mo día trabajado en la semana
        _verificar_septimo_dia(empleado, hoy, horas_netas_hoy)

    # ── DETECCIÓN DE PUNTUALIDAD Y CREACIÓN DE ALERTAS INTERNAS ─────────────
    _evaluar_alertas_asistencia(registro, empleado, registros_actualizados, horas_netas_hoy, es_offline=False)

    horas_restantes_hoy = max(0.0, round(8.0 - horas_netas_hoy, 2))
    cumplio_meta = horas_netas_hoy >= 7.95

    # Detectar si entró en el turno vespertino especial de las 5:00 PM (ej. Xiomara Castillo)
    primera_ent_kiosco = next((r for r in registros_actualizados if r.tipo_evento == 'ENTRADA'), None)
    es_turno_5pm_kiosco = False
    if primera_ent_kiosco:
        dt_pe = primera_ent_kiosco.fecha_hora.astimezone(timezone.get_current_timezone())
        mins_pe = dt_pe.hour * 60 + dt_pe.minute
        if abs(mins_pe - 1020) <= 45:
            es_turno_5pm_kiosco = True

    if tipo_evento == 'ENTRADA':
        mensaje_kiosco = f"¡Hola {empleado.nombre}! Entrada registrada. Jornada iniciada."
    elif tipo_evento == 'SALIDA_QUEBRADA':
        mensaje_kiosco = f"¡Buen descanso, {empleado.nombre}! Llevas {round(horas_netas_hoy, 1)} hrs acumuladas. Te faltan {round(horas_restantes_hoy, 1)} hrs para tus 8h."
    elif tipo_evento == 'ENTRADA_QUEBRADA':
        mensaje_kiosco = f"¡Bienvenido de vuelta, {empleado.nombre}! Llevas {round(horas_netas_hoy, 1)} hrs del primer turno. Te restan {round(horas_restantes_hoy, 1)} hrs para tus 8h."
    elif tipo_evento == 'SALIDA_DEFINITIVA':
        if comp_info['horas_amortizadas'] > 0:
            if comp_info['deuda_restante'] <= 0:
                if comp_info['horas_extra_solicitadas'] > 0:
                    mensaje_kiosco = (
                        f"¡Excelente trabajo, {empleado.nombre}! Completaste {round(horas_netas_hoy, 1)} hrs hoy. "
                        f"¡Saldaste tus {round(comp_info['horas_amortizadas'], 1)} hrs debidas 🎉 y enviamos {round(comp_info['horas_extra_solicitadas'], 1)} hrs extra para aprobación!"
                    )
                else:
                    mensaje_kiosco = (
                        f"¡Excelente trabajo, {empleado.nombre}! Completaste {round(horas_netas_hoy, 1)} hrs hoy. "
                        f"¡Has liquidado por completo tu deuda de horas pendientes ({round(comp_info['horas_amortizadas'], 1)} hrs saldadas)! 🎉 ¡Buen descanso!"
                    )
            else:
                mensaje_kiosco = (
                    f"Jornada finalizada, {empleado.nombre}. Completaste {round(horas_netas_hoy, 1)} hrs hoy. "
                    f"Se abonaron {round(comp_info['horas_amortizadas'], 1)} hrs a tu déficit de horas (Saldo pendiente restante: {round(comp_info['deuda_restante'], 1)} hrs)."
                )
        elif cumplio_meta:
            if comp_info['horas_extra_solicitadas'] > 0:
                mensaje_kiosco = (
                    f"¡Excelente trabajo, {empleado.nombre}! Completaste {round(horas_netas_hoy, 1)} hrs hoy (Meta de 8 hrs cumplida 🎉). "
                    f"Se enviaron {round(comp_info['horas_extra_solicitadas'], 1)} hrs extra para aprobación."
                )
            else:
                mensaje_kiosco = f"¡Excelente trabajo, {empleado.nombre}! Completaste {round(horas_netas_hoy, 1)} hrs hoy (Meta de 8 hrs cumplida 🎉). ¡Buen descanso!"
        elif es_turno_5pm_kiosco and horas_netas_hoy >= 5.8:
            mensaje_kiosco = (
                f"¡Buen trabajo, {empleado.nombre}! Turno vespertino completado ({round(horas_netas_hoy, 1)} hrs). "
                f"Déficit diario de {round(horas_restantes_hoy, 1)} hrs acumulado para reposición el fin de semana. ¡Buen descanso!"
            )
        else:
            mensaje_kiosco = f"Jornada finalizada, {empleado.nombre}. Acumulaste {round(horas_netas_hoy, 1)} hrs hoy (Déficit de {round(horas_restantes_hoy, 1)} hrs)."
    else:
        mensaje_kiosco = f"¡Hola {empleado.nombre}! Marcaje registrado correctamente."

    return Response({
        'status': 'ok',
        'mensaje': mensaje_kiosco,
        'registro': RegistroAsistenciaSerializer(registro, context={'request': request}).data,
        'horas_trabajadas_hoy': round(horas_netas_hoy, 2),
        'horas_restantes_hoy': round(horas_restantes_hoy, 2),
        'cumplio_meta_8h': cumplio_meta,
        'compensacion': comp_info,
    })


@api_view(['POST'])
@permission_classes([permissions.AllowAny])
@renderer_classes([JSONRenderer])
def consultar_horas_kiosco(request):
    """
    Endpoint seguro de solo lectura para consulta de horas del colaborador desde el Kiosco.
    No entrega datos salariales ni contraseñas.
    """
    qr_token = str(request.data.get('qr_token', '')).strip()
    if not qr_token:
        return Response({'status': 'error', 'mensaje': 'Código QR no proporcionado.'}, status=400)

    # Buscar empleado por token exacto o por prefijo UUID
    empleado = Empleado.objects.filter(qr_code_token=qr_token, activo=True).first()
    if not empleado:
        raw_uuid = qr_token.split('.', 1)[0]
        empleado = Empleado.objects.filter(qr_code_token__startswith=raw_uuid, activo=True).first()

    if not empleado:
        return Response({'status': 'error', 'mensaje': 'Carnet de colaborador no reconocido o inactivo.'}, status=404)

    tz_ni = timezone.get_current_timezone()
    ahora = timezone.now().astimezone(tz_ni)
    hoy = ahora.date()

    # 1. Marcajes de hoy
    import datetime
    start_dt = timezone.make_aware(datetime.datetime.combine(hoy, datetime.time.min), tz_ni)
    end_dt = timezone.make_aware(datetime.datetime.combine(hoy, datetime.time.max), tz_ni)

    registros_hoy = list(RegistroAsistencia.objects.filter(
        empleado=empleado,
        fecha_hora__range=(start_dt, end_dt)
    ).order_by('fecha_hora'))

    marcajes_hoy = [
        {
            'tipo': r.tipo_evento,
            'display': r.get_tipo_evento_display(),
            'hora': r.fecha_hora.astimezone(tz_ni).strftime('%I:%M %p')
        }
        for r in registros_hoy
    ]

    horas_hoy = round(_calcular_horas_netas_dia(registros_hoy), 2)

    # 2. Horas acumuladas del mes vigente
    primer_dia_mes = hoy.replace(day=1)
    start_mes = timezone.make_aware(datetime.datetime.combine(primer_dia_mes, datetime.time.min), tz_ni)

    registros_mes = list(RegistroAsistencia.objects.filter(
        empleado=empleado,
        fecha_hora__range=(start_mes, end_dt)
    ).order_by('fecha_hora'))

    por_dia = {}
    for r in registros_mes:
        d = r.fecha_hora.astimezone(tz_ni).date()
        por_dia.setdefault(d, []).append(r)

    horas_mes_total = sum(_calcular_horas_netas_dia(regs_dia) for regs_dia in por_dia.values())

    # 3. Estado de Turno en Curso (En Vivo)
    ultimo_reg = registros_hoy[-1] if registros_hoy else None
    turno_activo = False
    hora_inicio_turno = None
    horas_en_curso = 0.0

    if ultimo_reg and ultimo_reg.tipo_evento in ['ENTRADA', 'ENTRADA_QUEBRADA']:
        turno_activo = True
        hora_inicio_turno = ultimo_reg.fecha_hora.astimezone(tz_ni).strftime('%I:%M %p')
        segundos_transcurridos = max(0, (ahora - ultimo_reg.fecha_hora).total_seconds())
        horas_en_curso = round(segundos_transcurridos / 3600.0, 1)

    # 4. Horas Extra Aprobadas y Pendientes del Mes
    from django.db.models import Sum
    horas_extra_aprobadas = AutorizacionHorasExtra.objects.filter(
        empleado=empleado,
        estado='APROBADO',
        fecha__gte=primer_dia_mes,
        fecha__lte=hoy
    ).aggregate(total=Sum('horas_extra_autorizadas'))['total'] or 0.0

    horas_extra_pendientes = AutorizacionHorasExtra.objects.filter(
        empleado=empleado,
        estado='PENDIENTE',
        fecha__gte=primer_dia_mes,
        fecha__lte=hoy
    ).aggregate(total=Sum('horas_extra_solicitadas'))['total'] or 0.0

    dias_trabajados_mes = len(por_dia)

    return Response({
        'status': 'ok',
        'empleado': {
            'id': empleado.id,
            'nombre': empleado.nombre,
            'apellido': empleado.apellido,
            'cargo_display': empleado.get_cargo_display(),
        },
        'marcajes_hoy': marcajes_hoy,
        'horas_trabajadas_hoy': horas_hoy,
        'horas_mes': round(horas_mes_total, 1),
        'horas_pendientes': round(float(empleado.horas_pendientes), 2),
        'turno_activo': turno_activo,
        'hora_inicio_turno': hora_inicio_turno,
        'horas_en_curso': horas_en_curso,
        'horas_extra_aprobadas': round(float(horas_extra_aprobadas), 1),
        'horas_extra_pendientes': round(float(horas_extra_pendientes), 1),
        'dias_trabajados_mes': dias_trabajados_mes,
    })


def _calcular_horas_netas_dia(registros_dia):
    """
    Suma el tiempo entre ENTRADA -> SALIDA_QUEBRADA y ENTRADA_QUEBRADA -> SALIDA_DEFINITIVA.
    Aplica la Regla de Inicio Oficial (Opción A):
    Si un empleado llega temprano (ej. 11:26 AM para turno de 12:00 PM, o 5:53 PM para retorno de 6:00 PM),
    la jornada cuenta desde la hora oficial del turno, evitando inflar horas extras no autorizadas.
    """
    import datetime
    total_segundos = 0
    entrada_temp = None

    tz_ni = timezone.get_current_timezone()

    # Detectar si el día tiene quebrado y la hora de retorno
    retorno_q = next((r for r in registros_dia if r.tipo_evento == 'ENTRADA_QUEBRADA'), None)
    es_quebrado_11am = False
    if retorno_q:
        dt_ret_local = retorno_q.fecha_hora.astimezone(tz_ni)
        mins_ret = dt_ret_local.hour * 60 + dt_ret_local.minute
        if mins_ret >= 1115:  # 6:35 PM o después -> turno de 11:00 AM (retorno 7 PM)
            es_quebrado_11am = True

    for reg in registros_dia:
        dt_local = reg.fecha_hora.astimezone(tz_ni)

        if reg.tipo_evento == 'ENTRADA':
            mins_e = dt_local.hour * 60 + dt_local.minute
            # Determinar franja base oficial
            if es_quebrado_11am:
                base_mins = 660  # 11:00 AM
            elif 675 <= mins_e <= 735:
                # Entre 11:15 AM y 12:15 PM -> Turno oficial 12:00 PM (ej. Uriel 11:26 AM o Carlos 11:55 AM)
                base_mins = 720  # 12:00 PM
            else:
                FRANJAS = [540, 660, 720, 900, 1020]
                base_mins = min(FRANJAS, key=lambda b: abs(mins_e - b))

            h_b = base_mins // 60
            m_b = base_mins % 60
            inicio_oficial = timezone.make_aware(
                datetime.datetime.combine(dt_local.date(), datetime.time(h_b, m_b)),
                tz_ni
            )

            # Opción A: Si llegó antes de la hora oficial (hasta 60 min antes), arranca a la hora oficial
            if dt_local < inicio_oficial and (inicio_oficial - dt_local).total_seconds() <= 3600:
                entrada_temp = inicio_oficial
            else:
                entrada_temp = reg.fecha_hora

        elif reg.tipo_evento == 'ENTRADA_QUEBRADA':
            hora_ret_esperada = 1140 if es_quebrado_11am else 1080  # 7:00 PM o 6:00 PM
            h_b = hora_ret_esperada // 60
            m_b = hora_ret_esperada % 60
            retorno_oficial = timezone.make_aware(
                datetime.datetime.combine(dt_local.date(), datetime.time(h_b, m_b)),
                tz_ni
            )

            # Opción A: Si regresó antes de la hora oficial (hasta 45 min antes), arranca a la hora oficial
            if dt_local < retorno_oficial and (retorno_oficial - dt_local).total_seconds() <= 2700:
                entrada_temp = retorno_oficial
            else:
                entrada_temp = reg.fecha_hora

        elif reg.tipo_evento in ('SALIDA_QUEBRADA', 'SALIDA_DEFINITIVA') and entrada_temp:
            diferencia = (reg.fecha_hora - entrada_temp).total_seconds()
            if diferencia > 0:
                total_segundos += diferencia
            entrada_temp = None

    return total_segundos / 3600.0


def _recalcular_horas_pendientes_empleado(empleado, fecha_referencia=None):
    """
    Recalcula de forma determinista y auditable el saldo de horas_pendientes del empleado
    para el mes correspondiente a fecha_referencia (o el mes actual por defecto).
    Recorre cronológicamente los días que cuentan con SALIDA_DEFINITIVA registrada.
    Si se eliminó una salida definitiva o se corrigió un marcaje por error humano,
    este recálculo elimina deudas huérfanas y garantiza consistencia absoluta.
    """
    import datetime
    tz_ni = timezone.get_current_timezone()
    hoy = fecha_referencia or timezone.localdate()
    primer_dia_mes = hoy.replace(day=1)

    start_mes = timezone.make_aware(datetime.datetime.combine(primer_dia_mes, datetime.time.min), tz_ni)
    end_mes = timezone.make_aware(datetime.datetime.combine(hoy, datetime.time.max), tz_ni)

    registros_mes = list(RegistroAsistencia.objects.filter(
        empleado=empleado,
        fecha_hora__range=(start_mes, end_mes)
    ).order_by('fecha_hora'))

    por_dia = {}
    for r in registros_mes:
        d = r.fecha_hora.astimezone(tz_ni).date()
        por_dia.setdefault(d, []).append(r)

    deuda_acumulada = 0.0

    for d in sorted(por_dia.keys()):
        regs_d = por_dia[d]
        tiene_salida_definitiva = any(r.tipo_evento == 'SALIDA_DEFINITIVA' for r in regs_d)
        
        # Solo calculamos balance si la jornada del día está cerrada con salida definitiva
        if tiene_salida_definitiva:
            horas_dia = _calcular_horas_netas_dia(regs_d)
            if horas_dia < 8.0:
                deficit = round(8.0 - horas_dia, 2)
                deuda_acumulada += deficit
    # Restar compensaciones formalmente aprobadas en el mes
    from django.db.models import Sum
    total_compensado = CompensacionHoras.objects.filter(
        empleado=empleado,
        fecha_compensacion__range=(primer_dia_mes, hoy)
    ).aggregate(t=Sum('horas_deducidas'))['t'] or 0.0

    deuda_neta = max(0.0, round(deuda_acumulada - float(total_compensado), 2))
    empleado.horas_pendientes = deuda_neta
    empleado.periodo_horas_pendientes = primer_dia_mes
    empleado.save(update_fields=['horas_pendientes', 'periodo_horas_pendientes'])
    return empleado.horas_pendientes


def _acumular_horas_pendientes(empleado, fecha_hoy, horas_trabajadas_dia):
    """
    Si el empleado trabajó menos de 8 horas, acumula el déficit en horas_pendientes.
    Reinicia el saldo si el mes cambió desde el último registro del período.
    """
    horas_ordinarias = min(horas_trabajadas_dia, 8.0)
    deficit = max(0.0, 8.0 - horas_ordinarias)

    if deficit <= 0:
        return  # No hay deuda que acumular este día

    primer_dia_mes = fecha_hoy.replace(day=1)

    # Reiniciar si el período cambió (nuevo mes)
    if empleado.periodo_horas_pendientes != primer_dia_mes:
        empleado.horas_pendientes = 0.00
        empleado.periodo_horas_pendientes = primer_dia_mes

    empleado.horas_pendientes = float(empleado.horas_pendientes) + round(deficit, 2)
    empleado.save(update_fields=['horas_pendientes', 'periodo_horas_pendientes'])


def _aplicar_amortizacion_deuda_empleado(empleado, fecha_referencia, horas_a_amortizar, request=None):
    """
    Aplica formalmente la amortización de deuda acumulada (Bolsa de Horas)
    cuando el Administrador APRUEBA una solicitud de horas extra con PIN 2322.
    - Descuenta hasta min(horas_a_amortizar, deuda_actual) de empleado.horas_pendientes.
    - Genera registro auditable en CompensacionHoras con desglose FIFO de días adeudados.
    - Genera notificación en AlertaAsistencia(tipo='COMPENSACION_HORAS').
    - Registra en BitacoraAccion.
    """
    import datetime
    tz_ni = timezone.get_current_timezone()
    primer_dia_mes = fecha_referencia.replace(day=1)

    if empleado.periodo_horas_pendientes != primer_dia_mes:
        empleado.horas_pendientes = 0.00
        empleado.periodo_horas_pendientes = primer_dia_mes
        empleado.save(update_fields=['horas_pendientes', 'periodo_horas_pendientes'])

    deuda_actual = float(empleado.horas_pendientes or 0.0)
    if deuda_actual <= 0 or horas_a_amortizar <= 0:
        return {
            'horas_amortizadas': 0.0,
            'deuda_previa': deuda_actual,
            'deuda_restante': deuda_actual,
            'remanente': horas_a_amortizar,
            'compensacion': None,
        }

    horas_amortizadas = min(horas_a_amortizar, deuda_actual)
    nueva_deuda = max(0.0, round(deuda_actual - horas_amortizadas, 2))
    empleado.horas_pendientes = nueva_deuda
    empleado.save(update_fields=['horas_pendientes', 'periodo_horas_pendientes'])

    remanente_extra = max(0.0, round(horas_a_amortizar - horas_amortizadas, 2))

    # Buscar días previos con déficit para construir el desglose detallado (FIFO)
    start_periodo = timezone.make_aware(datetime.datetime.combine(primer_dia_mes, datetime.time.min), tz_ni)
    end_ayer = timezone.make_aware(datetime.datetime.combine(fecha_referencia - datetime.timedelta(days=1), datetime.time.max), tz_ni)

    regs_periodo = RegistroAsistencia.objects.filter(
        empleado=empleado,
        fecha_hora__range=(start_periodo, end_ayer)
    ).order_by('fecha_hora')

    dias_anteriores = {}
    for r in regs_periodo:
        dia_local = r.fecha_hora.astimezone(tz_ni).date()
        dias_anteriores.setdefault(dia_local, []).append(r)

    dias_deficit = []
    for dia_k, regs_k in sorted(dias_anteriores.items()):
        h_dia = _calcular_horas_netas_dia(regs_k)
        if 0 < h_dia < 8.0:
            deficit_k = round(8.0 - h_dia, 2)
            dias_deficit.append({
                'fecha': dia_k,
                'horas_trabajadas': round(h_dia, 2),
                'horas_faltaron': deficit_k,
            })

    desglose = []
    bolsa_disponible = horas_amortizadas

    for dd in dias_deficit:
        if bolsa_disponible <= 0:
            break
        faltan = dd['horas_faltaron']
        aplicadas = min(faltan, bolsa_disponible)
        saldo_dia = round(faltan - aplicadas, 2)
        estado_dia = "Liquidada al 100%" if saldo_dia == 0 else f"Abonada ({saldo_dia} hrs pendientes)"

        desglose.append({
            'fecha': dd['fecha'].strftime('%Y-%m-%d'),
            'fecha_display': dd['fecha'].strftime('%d/%m/%Y'),
            'horas_trabajadas': dd['horas_trabajadas'],
            'horas_faltaron': faltan,
            'horas_aplicadas': round(aplicadas, 2),
            'saldo_dia': saldo_dia,
            'estado': estado_dia,
        })
        bolsa_disponible = round(bolsa_disponible - aplicadas, 2)

    if not desglose and horas_amortizadas > 0:
        desglose.append({
            'fecha': primer_dia_mes.strftime('%Y-%m-%d'),
            'fecha_display': f"Período {primer_dia_mes.strftime('%m/%Y')}",
            'horas_trabajadas': round(8.0 - deuda_actual, 2) if deuda_actual < 8.0 else 0.0,
            'horas_faltaron': deuda_actual,
            'horas_aplicadas': horas_amortizadas,
            'saldo_dia': nueva_deuda,
            'estado': "Liquidada al 100%" if nueva_deuda == 0 else f"Abonada ({nueva_deuda} hrs pendientes)",
        })

    # Calcular horas trabajadas en la fecha de referencia
    start_dia_ref = timezone.make_aware(datetime.datetime.combine(fecha_referencia, datetime.time.min), tz_ni)
    end_dia_ref = timezone.make_aware(datetime.datetime.combine(fecha_referencia, datetime.time.max), tz_ni)
    regs_ref = list(RegistroAsistencia.objects.filter(
        empleado=empleado,
        fecha_hora__range=(start_dia_ref, end_dia_ref)
    ).order_by('fecha_hora'))
    horas_dia_ref = _calcular_horas_netas_dia(regs_ref) if regs_ref else round(8.0 + horas_a_amortizar, 2)

    # Registrar CompensacionHoras auditable
    comp = CompensacionHoras.objects.create(
        empleado=empleado,
        fecha_compensacion=fecha_referencia,
        horas_trabajadas_hoy=round(horas_dia_ref, 2),
        horas_extra_generadas=round(horas_a_amortizar, 2),
        horas_deducidas=round(horas_amortizadas, 2),
        deuda_previa=round(deuda_actual, 2),
        saldo_restante=round(nueva_deuda, 2),
        remanente_extra=round(remanente_extra, 2),
        desglose=desglose,
    )

    # Crear notificación en campanita para el Administrador
    AlertaAsistencia.objects.create(
        tipo='COMPENSACION_HORAS',
        empleado=empleado,
        titulo=f"Compensación Autorizada: {empleado.nombre} {empleado.apellido}",
        mensaje=(
            f"Al autorizar horas extra del {fecha_referencia.strftime('%d/%m/%Y')}, "
            f"se aplicaron {round(horas_amortizadas, 2)} hrs extra para amortizar su déficit acumulado. "
            f"Saldo de horas debidas actualizado a: {round(nueva_deuda, 2)} hrs."
        ),
        leida=False
    )

    # Registrar en bitácora auditable
    BitacoraAccion.objects.create(
        usuario=request.user if (request and hasattr(request, 'user') and request.user.is_authenticated) else None,
        accion='REGISTRO_MANUAL',
        descripcion=(
            f"Bolsa de Horas Aprobada: {round(horas_amortizadas, 2)} hrs compensadas para "
            f"{empleado.nombre} {empleado.apellido}. Deuda actualizada a {round(nueva_deuda, 2)} hrs."
        ),
        ip_address=_get_clean_ip(request) if request else None
    )

    return {
        'horas_amortizadas': round(horas_amortizadas, 2),
        'deuda_previa': round(deuda_actual, 2),
        'deuda_restante': round(nueva_deuda, 2),
        'remanente': round(remanente_extra, 2),
        'compensacion': comp,
    }


def _procesar_compensacion_y_horas_extra(empleado, fecha_hoy, horas_trabajadas_dia, request=None):
    """
    Gestiona la acumulación de déficit o creación de solicitud de Horas Extra al marcar salida:
    1. Si horas_trabajadas_dia < 8.0:
       - Acumula déficit en horas_pendientes del empleado.
    2. Si horas_trabajadas_dia >= 8.0:
       - Calcula excedente = horas_trabajadas_dia - 8.0.
       - NO amortiza deudas en el Kiosco. El excedente completo se envía como solicitud
         de Horas Extra en estado PENDIENTE.
       - Si y solo si el Administrador APRUEBA formalmente la solicitud con PIN 2322,
         se aplicará la deducción para amortizar la deuda acumulada del colaborador.
    """
    primer_dia_mes = fecha_hoy.replace(day=1)
    if empleado.periodo_horas_pendientes != primer_dia_mes:
        empleado.horas_pendientes = 0.00
        empleado.periodo_horas_pendientes = primer_dia_mes
        empleado.save(update_fields=['horas_pendientes', 'periodo_horas_pendientes'])

    if horas_trabajadas_dia < 8.0:
        _acumular_horas_pendientes(empleado, fecha_hoy, horas_trabajadas_dia)
        return {
            'horas_netas': horas_trabajadas_dia,
            'excedente': 0.0,
            'horas_amortizadas': 0.0,
            'deuda_restante': float(empleado.horas_pendientes),
            'horas_extra_solicitadas': 0.0,
        }

    excedente = round(horas_trabajadas_dia - 8.0, 2)
    deuda_actual = float(empleado.horas_pendientes)

    # El excedente pasa a solicitud PENDIENTE para revisión y aprobación de gerencia
    if excedente > 0.05:
        AutorizacionHorasExtra.objects.update_or_create(
            empleado=empleado,
            fecha=fecha_hoy,
            defaults={
                'horas_extra_solicitadas': round(excedente, 2),
                'estado': 'PENDIENTE'
            }
        )
    else:
        AutorizacionHorasExtra.objects.filter(
            empleado=empleado,
            fecha=fecha_hoy,
            estado='PENDIENTE'
        ).delete()

    return {
        'horas_netas': horas_trabajadas_dia,
        'excedente': excedente,
        'horas_amortizadas': 0.0,
        'deuda_restante': deuda_actual,
        'horas_extra_solicitadas': excedente,
    }



def _verificar_septimo_dia(empleado, fecha_hoy, horas_trabajadas_dia):
    """
    Detecta si el empleado ya trabajó 6 días anteriores en la misma semana ISO.
    Si es así, las horas del día actual se generan como solicitud de horas extra.
    """
    import datetime
    # Semana ISO: Lunes=0, Domingo=6
    inicio_semana = fecha_hoy - datetime.timedelta(days=fecha_hoy.weekday())
    fin_semana = inicio_semana + datetime.timedelta(days=6)

    # Contar días distintos con al menos una ENTRADA en la semana (excluyendo hoy)
    dias_trabajados = RegistroAsistencia.objects.filter(
        empleado=empleado,
        tipo_evento='ENTRADA',
        fecha_hora__date__gte=inicio_semana,
        fecha_hora__date__lt=fecha_hoy,
    ).dates('fecha_hora', 'day').count()

    if dias_trabajados >= 6:
        # El empleado ya cumplió sus 6 días — hoy es el 7mo, todo va a extra
        horas_extra_7mo = min(horas_trabajadas_dia, 8.0)
        if horas_extra_7mo > 0:
            AutorizacionHorasExtra.objects.update_or_create(
                empleado=empleado,
                fecha=fecha_hoy,
                defaults={
                    'horas_extra_solicitadas': round(horas_extra_7mo, 2),
                    'estado': 'PENDIENTE'
                }
            )


# ==========================================
# 📊 GENERADOR DE REPORTES EXCEL DE NÓMINA
# ==========================================

@api_view(['GET'])
@permission_classes([permissions.AllowAny])
@renderer_classes([ExcelBinaryRenderer, JSONRenderer])
def exportar_reporte_nomina_excel(request):
    """
    Genera una hoja de cálculo Excel (.xlsx) con el resumen de horas ordinarias y extras autorizadas.
    """
    try:
        import datetime
        from datetime import datetime as dt
        inicio_str = request.GET.get('fecha_inicio')
        fin_str = request.GET.get('fecha_fin')

        hoy = timezone.localdate()
        fecha_inicio = dt.strptime(inicio_str, '%Y-%m-%d').date() if inicio_str else hoy.replace(day=1)
        fecha_fin = dt.strptime(fin_str, '%Y-%m-%d').date() if fin_str else hoy

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BodegónPass"
        ws.views.sheetView[0].showGridLines = True

        # Estilos de Excel (Basado en el Verde Institucional de El Bodegón)
        COLOR_HEADER = "1C6856"
        COLOR_TITLE = "134F42"
        font_titulo = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
        font_sub = Font(name='Calibri', size=11, italic=True, color='FFFFFF')
        font_header = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        font_data = Font(name='Calibri', size=11)
        font_bold = Font(name='Calibri', size=11, bold=True)

        fill_header = PatternFill(fill_type='solid', start_color=COLOR_HEADER, end_color=COLOR_HEADER)
        fill_title = PatternFill(fill_type='solid', start_color=COLOR_TITLE, end_color=COLOR_TITLE)
        fill_zebra = PatternFill(fill_type='solid', start_color='F9F9F9', end_color='F9F9F9')

        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

        # Headers — 7 Columnas Ejecutivas
        headers = [
            "Empleado y Puesto",
            "Días Trabajados",
            "Días Libres (Tomados)",
            "Horas Ordinarias",
            "Horas Feriados (Días)",
            "Horas Extra Aprobadas",
            "Horas Debidas (Déficit)",
        ]

        ws.merge_cells('A1:G1')
        ws['A1'] = "BODEGÓN PASS — REPORTE DE ASISTENCIA Y PERSONAL"
        ws['A1'].font = font_titulo
        ws['A1'].fill = fill_title
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('A2:G2')
        ws['A2'] = f"Período del {fecha_inicio.strftime('%d/%m/%Y')} al {fecha_fin.strftime('%d/%m/%Y')} — Generado el {hoy.strftime('%d/%m/%Y')}"
        ws['A2'].font = font_sub
        ws['A2'].fill = fill_title
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

        ws.append([])        # Fila 3 vacía
        ws.append(headers)   # Fila 4 Headers

        for col in range(1, 8):
            cell = ws.cell(row=4, column=col)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal='center', vertical='center')

        empleados = Empleado.objects.filter(activo=True)
        row_idx = 5

        # Obtener feriados en el rango
        feriados_set = set(DiaFeriado.objects.filter(
            fecha__gte=fecha_inicio,
            fecha__lte=fecha_fin
        ).values_list('fecha', flat=True))

        # Obtener permisos y vacaciones autorizadas en el rango
        permisos_qs = PermisoAusencia.objects.filter(
            fecha_inicio__lte=fecha_fin,
            fecha_fin__gte=fecha_inicio
        ).select_related('empleado')

        permisos_por_empleado = {}
        permisos_info_por_empleado = {}
        for p in permisos_qs:
            if p.empleado_id not in permisos_por_empleado:
                permisos_por_empleado[p.empleado_id] = set()
                permisos_info_por_empleado[p.empleado_id] = []
            curr_p = max(p.fecha_inicio, fecha_inicio)
            fin_p = min(p.fecha_fin, fecha_fin)
            permisos_info_por_empleado[p.empleado_id].append(
                f"{p.get_tipo_display()} ({curr_p.strftime('%d/%m')} a {fin_p.strftime('%d/%m')})"
            )
            while curr_p <= fin_p:
                permisos_por_empleado[p.empleado_id].add(curr_p)
                curr_p += datetime.timedelta(days=1)

        for emp in empleados:
            start_query = timezone.make_aware(datetime.datetime.combine(fecha_inicio, datetime.time.min), timezone.get_current_timezone())
            end_query = timezone.make_aware(datetime.datetime.combine(fecha_fin, datetime.time.max), timezone.get_current_timezone())
            
            registros = RegistroAsistencia.objects.filter(
                empleado=emp,
                fecha_hora__range=(start_query, end_query)
            ).order_by('fecha_hora')

            # Agrupar registros por día local (Nicaragua)
            horas_normales_trabajadas = 0.0
            feriados_trabajados_dias = 0
            feriados_trabajados_info = []
            dias_map = {}
            for reg in registros:
                reg_local = reg.fecha_hora.astimezone(timezone.get_current_timezone())
                dia_local = reg_local.date()
                if dia_local < fecha_inicio or dia_local > fecha_fin:
                    continue
                if dia_local not in dias_map:
                    dias_map[dia_local] = []
                dias_map[dia_local].append(reg)

            dias_trabajados = len(dias_map)
            dias_permiso_emp = permisos_por_empleado.get(emp.id, set())

            for dia, regs in dias_map.items():
                horas_dia = _calcular_horas_netas_dia(regs)
                horas_ord = min(horas_dia, 8.0)
                if dia in feriados_set:
                    if horas_ord >= 8.0:
                        feriados_trabajados_dias += 1
                        desc_feriado = DiaFeriado.objects.filter(fecha=dia).first()
                        nombre_feriado = desc_feriado.descripcion if desc_feriado else "Día Feriado"
                        feriados_trabajados_info.append(
                            f"- {dia.strftime('%d/%m/%Y')}: {nombre_feriado} ({round(horas_ord, 2)} hrs)"
                        )
                    else:
                        horas_normales_trabajadas += horas_ord
                else:
                    horas_normales_trabajadas += horas_ord

            # ── Detectar días libres y ausencias extra usando lógica semanal ────
            # Por cada semana en el período, el primer día sin marcaje = día libre.
            # Los días con permiso/vacaciones autorizadas NO suman faltas ni deudas.
            dias_libres = 0
            horas_debidas = 0.0

            curr_day = fecha_inicio
            semanas_procesadas = set()

            while curr_day <= fecha_fin:
                semana_key = curr_day - datetime.timedelta(days=curr_day.weekday())
                if semana_key not in semanas_procesadas:
                    semanas_procesadas.add(semana_key)
                    semana_fin = semana_key + datetime.timedelta(days=6)
                    s_inicio = max(semana_key, fecha_inicio)
                    s_fin = min(semana_fin, fecha_fin)

                    ausencias_semana = 0
                    deficit_semana = 0.0
                    excedente_semana = 0.0
                    d = s_inicio
                    while d <= s_fin:
                        # Si es feriado o permiso autorizado, se exonera de falta y deuda
                        if d not in feriados_set and d not in dias_permiso_emp:
                            if d not in dias_map:
                                if ausencias_semana == 0:
                                    dias_libres += 1  # Primera ausencia = día libre
                                else:
                                    deficit_semana += 8.0   # Segunda+ = ausencia extra
                                ausencias_semana += 1
                            else:
                                # Día trabajado: calcular déficit y excedente
                                regs_d = dias_map[d]
                                horas_dia = _calcular_horas_netas_dia(regs_d)
                                horas_ord = min(horas_dia, 8.0)
                                deficit = max(0.0, 8.0 - horas_ord)
                                deficit_semana += deficit
                                if horas_dia > 8.0:
                                    excedente_semana += (horas_dia - 8.0)
                        d += datetime.timedelta(days=1)

                    # Compensar déficit de la semana con horas adicionales de la misma semana
                    compensado_semana = min(excedente_semana, deficit_semana)
                    deficit_neto_semana = max(0.0, deficit_semana - compensado_semana)
                    horas_debidas += deficit_neto_semana
                    # Las horas que compensaron deuda se suman a las horas ordinarias trabajadas
                    horas_normales_trabajadas += compensado_semana
                curr_day += datetime.timedelta(days=1)

            # Horas extra aprobadas
            from django.db.models import Sum
            horas_extra_aprobadas = AutorizacionHorasExtra.objects.filter(
                empleado=emp,
                fecha__gte=fecha_inicio,
                fecha__lte=fecha_fin,
                estado='APROBADO'
            ).aggregate(total=Sum('horas_extra_autorizadas'))['total'] or 0.0

            fila = [
                f"{emp.nombre} {emp.apellido} ({emp.get_cargo_display()})",
                dias_trabajados,
                dias_libres,
                round(horas_normales_trabajadas, 2),
                feriados_trabajados_dias,
                round(float(horas_extra_aprobadas), 2),
                round(horas_debidas, 2),
            ]
            ws.append(fila)

            from openpyxl.comments import Comment
            # Comentario de permisos/vacaciones en la celda del empleado
            info_permisos = permisos_info_por_empleado.get(emp.id)
            if info_permisos:
                cell_emp = ws.cell(row=row_idx, column=1)
                cell_emp.comment = Comment("Permisos/Vacaciones:\n" + "\n".join(info_permisos), "BodegónPass")

            if feriados_trabajados_dias > 0 and feriados_trabajados_info:
                comentario_texto = "Detalle de Feriados Laborados:\n" + "\n".join(feriados_trabajados_info)
                cell_feriado = ws.cell(row=row_idx, column=5)
                cell_feriado.comment = Comment(comentario_texto, "BodegónPass")

            for col in range(1, 8):
                cell = ws.cell(row=row_idx, column=col)
                cell.font = font_data
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = fill_zebra
                if col in [2, 3, 4, 5, 6, 7]:
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='left')

            row_idx += 1

        # Fila de Totales
        ws.append([])
        row_idx += 1
        ws.merge_cells(f'A{row_idx}:C{row_idx}')
        ws[f'A{row_idx}'] = "TOTALES GENERALES:"
        ws[f'A{row_idx}'].font = font_bold
        ws[f'A{row_idx}'].alignment = Alignment(horizontal='right')

        for col_letter in ['D', 'E', 'F', 'G']:
            cell = ws[f'{col_letter}{row_idx}']
            cell.value = f"=SUM({col_letter}5:{col_letter}{row_idx-2})"
            cell.font = font_bold
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='right')

        # Ajustar ancho de columnas
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        nombre_archivo = f"BodegonPass_Reporte_{fecha_inicio.strftime('%Y%m%d')}_{fecha_fin.strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        response['X-Filename'] = nombre_archivo
        response['Access-Control-Expose-Headers'] = 'Content-Disposition, X-Filename'

        BitacoraAccion.objects.create(
            usuario=request.user if request.user.is_authenticated else None,
            accion='EXPORTAR_NOMINA',
            descripcion=f"Reporte de horas y asistencia Excel exportado ({fecha_inicio} a {fecha_fin}).",
            ip_address=_get_clean_ip(request)
        )

        return response

    except Exception as e:
        return Response({'detail': f'Error generando reporte Excel: {str(e)}'}, status=500)


@api_view(['GET'])
@permission_classes([permissions.AllowAny])
@renderer_classes([ExcelBinaryRenderer, JSONRenderer])
def exportar_reporte_vacaciones_excel(request):
    """
    Genera un archivo Excel (.xlsx) mensual o por rango de fechas con el detalle completo de vacaciones y permisos:
    - Colaborador y Puesto
    - Tipo de Ausencia (Vacaciones, Vacaciones Pagadas, Incapacidad Médica, Permiso Autorizado)
    - Fechas de Inicio y Fin
    - Total de Días Concedidos
    - Motivo / Justificación
    - Fecha de Registro / Autorización
    """
    try:
        import datetime
        import calendar
        from datetime import datetime as dt
        from io import BytesIO

        mes_str = request.GET.get('mes')
        anio_str = request.GET.get('anio')
        inicio_str = request.GET.get('fecha_inicio')
        fin_str = request.GET.get('fecha_fin')

        hoy = timezone.localdate()

        MESES_ES = [
            '', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
            'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
        ]

        if mes_str and anio_str:
            mes = int(mes_str)
            anio = int(anio_str)
            _, ultimo_dia = calendar.monthrange(anio, mes)
            fecha_inicio = datetime.date(anio, mes, 1)
            fecha_fin = datetime.date(anio, mes, ultimo_dia)
            periodo_label = f"Mes de {MESES_ES[mes]} de {anio}"
        elif inicio_str and fin_str:
            fecha_inicio = dt.strptime(inicio_str, '%Y-%m-%d').date()
            fecha_fin = dt.strptime(fin_str, '%Y-%m-%d').date()
            periodo_label = f"Del {fecha_inicio.strftime('%d/%m/%Y')} al {fecha_fin.strftime('%d/%m/%Y')}"
        else:
            mes = hoy.month
            anio = hoy.year
            _, ultimo_dia = calendar.monthrange(anio, mes)
            fecha_inicio = datetime.date(anio, mes, 1)
            fecha_fin = datetime.date(anio, mes, ultimo_dia)
            periodo_label = f"Mes de {MESES_ES[mes]} de {anio}"

        # Filtrar permisos que caigan o se solapen con el período
        permisos = PermisoAusencia.objects.filter(
            fecha_inicio__lte=fecha_fin,
            fecha_fin__gte=fecha_inicio
        ).select_related('empleado').order_by('fecha_inicio', 'empleado__nombre')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vacaciones y Permisos"
        ws.views.sheetView[0].showGridLines = True

        # Colores Institucionales El Bodegón
        COLOR_HEADER = "1C6856"
        COLOR_TITLE = "134F42"
        COLOR_CARD = "F2F7F5"

        font_titulo = Font(name='Calibri', size=15, bold=True, color='FFFFFF')
        font_sub = Font(name='Calibri', size=10, italic=True, color='FFFFFF')
        font_header = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        font_data = Font(name='Calibri', size=10)
        font_bold = Font(name='Calibri', size=10, bold=True)
        font_mono = Font(name='Consolas', size=10)

        fill_title = PatternFill(fill_type='solid', start_color=COLOR_TITLE, end_color=COLOR_TITLE)
        fill_header = PatternFill(fill_type='solid', start_color=COLOR_HEADER, end_color=COLOR_HEADER)
        fill_zebra = PatternFill(fill_type='solid', start_color='F9FBF9', end_color='F9FBF9')
        fill_card = PatternFill(fill_type='solid', start_color=COLOR_CARD, end_color=COLOR_CARD)

        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

        # Fila 1: Título Principal
        ws.merge_cells('A1:I1')
        ws['A1'] = "EL BODEGÓN — REPORTE MENSUAL DE VACACIONES Y PERMISOS"
        ws['A1'].font = font_titulo
        ws['A1'].fill = fill_title
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 28

        # Fila 2: Subtítulo con Período
        ws.merge_cells('A2:I2')
        ws['A2'] = f"Período: {periodo_label} | Generado el {hoy.strftime('%d/%m/%Y')}"
        ws['A2'].font = font_sub
        ws['A2'].fill = fill_title
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 18

        # Tarjetas de Resumen (Fila 4)
        total_solicitudes = permisos.count()
        total_dias_otorgados = sum(p.total_dias for p in permisos)
        total_colaboradores_unicos = len(set(p.empleado_id for p in permisos))

        ws.merge_cells('A4:B4')
        ws['A4'] = f"Total Solicitudes: {total_solicitudes}"
        ws['A4'].font = font_bold
        ws['A4'].fill = fill_card
        ws['A4'].alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('D4:F4')
        ws['D4'] = f"Días Concedidos: {total_dias_otorgados} días"
        ws['D4'].font = font_bold
        ws['D4'].fill = fill_card
        ws['D4'].alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('H4:I4')
        ws['H4'] = f"Colaboradores: {total_colaboradores_unicos}"
        ws['H4'].font = font_bold
        ws['H4'].fill = fill_card
        ws['H4'].alignment = Alignment(horizontal='center', vertical='center')

        for cell_ref in ['A4', 'B4', 'D4', 'E4', 'F4', 'H4', 'I4']:
            ws[cell_ref].border = thin_border

        # Headers Tabla (Fila 6)
        headers = [
            "N°",
            "Colaborador",
            "Cargo",
            "Tipo de Ausencia",
            "Fecha Inicio",
            "Fecha Fin",
            "Días Concedidos",
            "Motivo / Justificación",
            "Fecha Registro",
        ]

        ws.row_dimensions[6].height = 24
        for col_idx, h in enumerate(headers, 1):
            c = ws.cell(row=6, column=col_idx, value=h)
            c.font = font_header
            c.fill = fill_header
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = thin_border

        # Filas de Datos
        row_idx = 7
        for i, p in enumerate(permisos, 1):
            ws.row_dimensions[row_idx].height = 20
            emp = p.empleado
            motivo_texto = p.motivo.strip() if p.motivo else f"Concesión de {p.get_tipo_display()}"
            fecha_reg_str = p.created_at.astimezone(timezone.get_current_timezone()).strftime('%d/%m/%Y %I:%M %p') if p.created_at else '-'

            row_data = [
                i,
                f"{emp.nombre} {emp.apellido}",
                emp.get_cargo_display() if hasattr(emp, 'get_cargo_display') else emp.cargo,
                p.get_tipo_display(),
                p.fecha_inicio.strftime('%d/%m/%Y'),
                p.fecha_fin.strftime('%d/%m/%Y'),
                p.total_dias,
                motivo_texto,
                fecha_reg_str,
            ]

            for col_idx, val in enumerate(row_data, 1):
                c = ws.cell(row=row_idx, column=col_idx, value=val)
                c.font = font_data
                c.border = thin_border
                if row_idx % 2 == 0:
                    c.fill = fill_zebra

                if col_idx in [1, 7]:
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    if col_idx == 7:
                        c.font = font_bold
                elif col_idx in [5, 6, 9]:
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    c.font = font_mono
                else:
                    c.alignment = Alignment(horizontal='left', vertical='center')

            row_idx += 1

        # Fila de Totales si hay registros
        if total_solicitudes > 0:
            ws.merge_cells(f'A{row_idx}:F{row_idx}')
            tot_label = ws.cell(row=row_idx, column=1, value="TOTAL DÍAS DE VACACIONES Y PERMISOS:")
            tot_label.font = font_bold
            tot_label.alignment = Alignment(horizontal='right', vertical='center')
            tot_label.fill = fill_card

            for c_i in range(1, 7):
                ws.cell(row=row_idx, column=c_i).border = thin_border
                ws.cell(row=row_idx, column=c_i).fill = fill_card

            tot_val = ws.cell(row=row_idx, column=7, value=f"=SUM(G7:G{row_idx-1})")
            tot_val.font = font_bold
            tot_val.alignment = Alignment(horizontal='center', vertical='center')
            tot_val.border = thin_border
            tot_val.fill = fill_card

            for c_i in [8, 9]:
                c_empty = ws.cell(row=row_idx, column=c_i, value="")
                c_empty.border = thin_border
                c_empty.fill = fill_card

            ws.row_dimensions[row_idx].height = 22
        else:
            ws.merge_cells(f'A{row_idx}:I{row_idx}')
            c_empty = ws.cell(row=row_idx, column=1, value="No se registraron vacaciones ni permisos en este período.")
            c_empty.font = font_data
            c_empty.alignment = Alignment(horizontal='center', vertical='center')
            c_empty.border = thin_border
            ws.row_dimensions[row_idx].height = 24

        # Auto-ajuste de columnas
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 24
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 16
        ws.column_dimensions['H'].width = 32
        ws.column_dimensions['I'].width = 20

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        nombre_archivo = f"BodegonPass_Vacaciones_{fecha_inicio.strftime('%Y%m')}.xlsx"
        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        response['X-Filename'] = nombre_archivo
        response['Access-Control-Expose-Headers'] = 'Content-Disposition, X-Filename'

        BitacoraAccion.objects.create(
            usuario=request.user if request.user.is_authenticated else None,
            accion='EXPORTAR_NOMINA',
            descripcion=f"Reporte mensual de vacaciones Excel exportado ({periodo_label}).",
            ip_address=_get_clean_ip(request)
        )

        return response
    except Exception as e:
        return Response({'detail': f'Error generando reporte mensual de vacaciones: {str(e)}'}, status=500)


@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def sync_batch_asistencia(request):
    """
    Sincroniza un lote de marcajes registrados sin conexión.
    """
    payload = request.data
    if not isinstance(payload, list):
        return Response({'detail': 'Se requiere una lista de marcajes.'}, status=400)

    # Ordenar cronológicamente por fecha_hora para garantizar que la auto-detección
    # de eventos se realice en la secuencia correcta.
    try:
        payload = sorted(payload, key=lambda x: x.get('fecha_hora', ''))
    except Exception as e:
        print(f"Error ordenando payload de sincronización: {e}")

    respuestas = []
    for item in payload:
        qr_code_token = item.get('qr_code_token')
        tipo_evento = item.get('tipo_evento')
        fecha_hora_str = item.get('fecha_hora')
        foto = item.get('foto')

        # 1. Validar HMAC
        import hmac
        import hashlib
        from django.conf import settings

        if not qr_code_token or '.' not in qr_code_token:
            respuestas.append({'status': 'error', 'mensaje': 'Formato QR inválido.', 'token': qr_code_token})
            continue

        # 1. Buscar Empleado por token o prefijo UUID
        empleado = Empleado.objects.filter(qr_code_token=qr_code_token.strip(), activo=True).first()
        if not empleado and '.' in qr_code_token:
            raw_uuid = qr_code_token.split('.', 1)[0]
            empleado = Empleado.objects.filter(qr_code_token__startswith=raw_uuid, activo=True).first()

        if not empleado:
            respuestas.append({'status': 'error', 'mensaje': 'Empleado no encontrado o inactivo.', 'token': qr_code_token})
            continue

        # 3. Guardar registro
        from django.core.files.base import ContentFile
        import base64
        from django.utils.dateparse import parse_datetime
        import datetime

        foto_file = None
        foto_b64 = None
        if foto and ',' in foto:
            foto_b64 = foto
            try:
                format, imgstr = foto.split(';base64,')
                ext = format.split('/')[-1]
                foto_file = ContentFile(base64.b64decode(imgstr), name=f"{empleado.id}_{int(timezone.now().timestamp())}.{ext}")
            except Exception as e:
                print(f"Error procesando base64: {e}")

        fecha_hora = parse_datetime(fecha_hora_str) if fecha_hora_str else timezone.now()
        # Asegurar zona horaria correcta
        if fecha_hora and timezone.is_naive(fecha_hora):
            fecha_hora = timezone.make_aware(fecha_hora, timezone.get_current_timezone())

        # Auto-detectar tipo_evento si no se proporciona
        if not tipo_evento:
            dt_local = fecha_hora.astimezone(timezone.get_current_timezone())
            dia = dt_local.date()
            start_dt = timezone.make_aware(datetime.combine(dia, datetime.time.min), timezone.get_current_timezone())
            end_dt = timezone.make_aware(datetime.combine(dia, datetime.time.max), timezone.get_current_timezone())
            
            registros_hoy = RegistroAsistencia.objects.filter(
                empleado=empleado,
                fecha_hora__range=(start_dt, end_dt)
            ).order_by('fecha_hora')

            if not tipo_evento or tipo_evento == 'AUTODETECT':
                tipo_evento = _autodetectar_tipo_evento(registros_hoy, fecha_hora)

        with transaction.atomic():
            registro = RegistroAsistencia.objects.create(
                empleado=empleado,
                tipo_evento=tipo_evento,
                foto_verificacion=foto_file,
                foto_base64=foto_b64,
                fecha_hora=fecha_hora,
                ip_address=_get_clean_ip(request)
            )

        # Calcular horas hoy y verificar tardanzas
        dt_local = fecha_hora.astimezone(timezone.get_current_timezone())
        dia = dt_local.date()
        
        import datetime
        start_dt = timezone.make_aware(datetime.datetime.combine(dia, datetime.time.min), timezone.get_current_timezone())
        end_dt = timezone.make_aware(datetime.datetime.combine(dia, datetime.time.max), timezone.get_current_timezone())
        
        registros_actualizados = list(RegistroAsistencia.objects.filter(
            empleado=empleado,
            fecha_hora__range=(start_dt, end_dt)
        ).order_by('fecha_hora'))

        # Lógica de horas extras y déficit
        horas_netas_hoy = _calcular_horas_netas_dia(registros_actualizados)

        # Si es ENTRADA, revisar si el día previo quedó sin salida
        if tipo_evento == 'ENTRADA':
            ayer = dia - datetime.timedelta(days=1)
            start_ayer = timezone.make_aware(datetime.datetime.combine(ayer, datetime.time.min), timezone.get_current_timezone())
            end_ayer = timezone.make_aware(datetime.datetime.combine(ayer, datetime.time.max), timezone.get_current_timezone())
            regs_ayer = list(RegistroAsistencia.objects.filter(
                empleado=empleado,
                fecha_hora__range=(start_ayer, end_ayer)
            ).order_by('fecha_hora'))
            if regs_ayer:
                ultimo_ayer = regs_ayer[-1].tipo_evento
                if ultimo_ayer in ('ENTRADA', 'ENTRADA_QUEBRADA'):
                    AlertaAsistencia.objects.get_or_create(
                        tipo='REGISTRO_INCOMPLETO',
                        empleado=empleado,
                        titulo=f"Registro incompleto (Offline): {empleado.nombre} {empleado.apellido}",
                        defaults={
                            'mensaje': (
                                f"El día {ayer.strftime('%d/%m/%Y')} el empleado registró "
                                f"{regs_ayer[-1].get_tipo_evento_display()} "
                                f"pero nunca registró su Salida. Por favor, agregue la salida manualmente."
                            ),
                            'leida': False
                        }
                    )

        if tipo_evento == 'SALIDA_DEFINITIVA':
            _procesar_compensacion_y_horas_extra(empleado, dia, horas_netas_hoy, request)
            # 7mo día de la semana
            _verificar_septimo_dia(empleado, dia, horas_netas_hoy)

        # Alertas de asistencia (Regla de gracia 10 min de El Bodegón)
        _evaluar_alertas_asistencia(registro, empleado, registros_actualizados, horas_netas_hoy, es_offline=True)

        respuestas.append({'status': 'ok', 'id': registro.id, 'empleado': empleado.nombre})

    return Response({'status': 'ok', 'sincronizados': respuestas})
